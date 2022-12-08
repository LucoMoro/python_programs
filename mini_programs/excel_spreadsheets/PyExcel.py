import openpyxl

#wb = openpyxl.Workbook()
wb = openpyxl.load_workbook("transactions.xlsx")
print(wb.sheetnames)

sheet = wb["Sheet1"]
#wb.create_sheet("Sheet2", 0)
#wb.remove_sheet(sheet)

#cell = sheet["a1"]
#column = sheet["a"]
#cells = sheet["a:c"]
sheet["a1:c3"]
sheet[1:3]
sheet.append([1, 2, 3])


for row in range(1, sheet.max_row + 1):
    for column in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=row, column=column)
        print(cell.value)

wb.save("transactions2.xlsx")